import os
import sys
import math
import datetime as dt
from collections import defaultdict, Counter

import pandas as pd

from PySide6 import QtCore, QtGui, QtWidgets

import tempfile
import shutil
from openpyxl import Workbook
import time


# =========================
# Jalali (Persian) Calendar
# =========================

def gregorian_to_jalali(gy: int, gm: int, gd: int):
    """
    Convert Gregorian date to Jalali (Persian) date.
    Returns (jy, jm, jd).
    Algorithm adapted from the public-domain JDF routines.
    """
    g_d_m = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334]
    if gy > 1600:
        jy = 979
        gy -= 1600
    else:
        jy = 0
        gy -= 621
    gy2 = gy + 1 if gm > 2 else gy
    days = (365 * gy) + ((gy2 + 3) // 4) - ((gy2 + 99) // 100) + ((gy2 + 399) // 400) - 80 + gd + g_d_m[gm - 1]
    jy += 33 * (days // 12053)
    days %= 12053
    jy += 4 * (days // 1461)
    days %= 1461
    if days > 365:
        jy += (days - 1) // 365
        days = (days - 1) % 365
    if days < 186:
        jm = 1 + (days // 31)
        jd = 1 + (days % 31)
    else:
        days -= 186
        jm = 7 + (days // 30)
        jd = 1 + (days % 30)
    return jy, jm, jd


def today_jalali_str(sep: str = "/") -> str:
    today = dt.date.today()
    jy, jm, jd = gregorian_to_jalali(today.year, today.month, today.day)
    return f"{jy:04d}{sep}{jm:02d}{sep}{jd:02d}"


# =========================
# Constants / Paths
# =========================

def app_dir():
    # When running as .exe (PyInstaller), sys._MEIPASS is temp; use exe folder instead.
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

NETWORK_DIR = r"\\192.168.20.15\DataCenter Office\IT"
EXCEL_PATH = os.path.join(NETWORK_DIR, "printers.xlsx")

SHEET_DATA = "Data"
SHEET_LOGS = "Logs"
SHEET_LISTS = "Lists"

DATA_COLUMNS = ["Row", "Model", "New in storage", "Storage", "Repair", "Total", "User"]
LOG_COLUMNS = ["Date (Jalali)", "Event", "User", "Model", "Quantity", "Notes"]
LISTS_COLUMNS = ["PrinterNames", "Usernames", "UserHasPrinters"]  # third col rows: "username|model"


# =========================
# Data Manager
# =========================

class DataManager(QtCore.QObject):
    data_changed = QtCore.Signal()

    def __init__(self, excel_path: str, parent=None):
        super().__init__(parent)
        self.excel_path = excel_path
        self._ensure_dir()
        self.data_df = None
        self.logs_df = None
        self.lists_df = None
        self.load()

    @staticmethod
    def _ensure_columns(df, columns):
        """
        Ensure df has all columns in `columns` (create missing ones),
        then return df ordered to exactly those columns.
        """
        for c in columns:
           if c not in df.columns:
                df[c] = pd.Series(dtype=object)
        return df[columns]


    def _ensure_dir(self):
        folder = os.path.dirname(self.excel_path)
        if folder and not os.path.exists(folder):
            os.makedirs(folder, exist_ok=True)

    # ---------- SAFE LOAD / INIT / SAVE ----------

    def _init_empty_workbook(self):
        """
        Create a valid Excel file via openpyxl (never corrupt),
        with 3 sheets and headers.
        """
        wb = Workbook()
        wb.active.title = SHEET_DATA
        wb.create_sheet(SHEET_LOGS)
        wb.create_sheet(SHEET_LISTS)

        ws_data = wb[SHEET_DATA]
        ws_data.append(DATA_COLUMNS)

        ws_logs = wb[SHEET_LOGS]
        ws_logs.append(LOG_COLUMNS)

        ws_lists = wb[SHEET_LISTS]
        ws_lists.append(LISTS_COLUMNS)

        wb.save(self.excel_path)

        # also init in-memory dfs
        self.data_df = pd.DataFrame(columns=DATA_COLUMNS)
        self.logs_df = pd.DataFrame(columns=LOG_COLUMNS)
        self.lists_df = pd.DataFrame(columns=LISTS_COLUMNS)

    def load(self):
        """
        Load all sheets. If file doesn't exist or is corrupt, rebuild it.
        """
        if not os.path.exists(self.excel_path):
            self._init_empty_workbook()

        try:
            dfs = pd.read_excel(self.excel_path, sheet_name=None, engine="openpyxl")
        except Exception:
            # file invalid/corrupt -> recreate
            self._init_empty_workbook()
            dfs = pd.read_excel(self.excel_path, sheet_name=None, engine="openpyxl")

        self.data_df = dfs.get(SHEET_DATA, pd.DataFrame(columns=DATA_COLUMNS))
        self.logs_df = dfs.get(SHEET_LOGS, pd.DataFrame(columns=LOG_COLUMNS))
        self.lists_df = dfs.get(SHEET_LISTS, pd.DataFrame(columns=LISTS_COLUMNS))

        self.data_df = self._ensure_columns(self.data_df, DATA_COLUMNS)
        self.logs_df = self._ensure_columns(self.logs_df, LOG_COLUMNS)
        self.lists_df = self._ensure_columns(self.lists_df, LISTS_COLUMNS)

        # numeric coercion
        for c in ["New in storage", "Storage", "Repair", "Total"]:
            self.data_df[c] = pd.to_numeric(self.data_df[c], errors="coerce").fillna(0).astype(int)

        # normalize Row
        if self.data_df.empty:
            self.data_df["Row"] = pd.Series(dtype=int)
        else:
            self.data_df["Row"] = range(1, len(self.data_df) + 1)

        self.sync_totals(save=False)

    def save(self):
        folder = os.path.dirname(self.excel_path)
        os.makedirs(folder, exist_ok=True)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=folder, delete=False) as tmp:
            tmp_path = tmp.name

        try:
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                self.data_df.to_excel(writer, sheet_name=SHEET_DATA, index=False)
                self.logs_df.to_excel(writer, sheet_name=SHEET_LOGS, index=False)
                self.lists_df.to_excel(writer, sheet_name=SHEET_LISTS, index=False)

            for attempt in range(5):
                try:
                    os.replace(tmp_path, self.excel_path)
                    break
                except PermissionError:
                    if attempt == 4:
                        raise
                    time.sleep(0.5)

        except Exception:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
            raise

    # ---------- LISTS (Sheet 3) ----------

    def get_printer_names(self) -> list:
        vals = self.lists_df["PrinterNames"].dropna().astype(str).tolist()
        seen, out = set(), []
        for v in vals:
            v = v.strip()
            if v and v not in seen:
                seen.add(v)
                out.append(v)
        return out

    def get_usernames(self) -> list:
        vals = self.lists_df["Usernames"].dropna().astype(str).tolist()
        seen, out = set(), []
        for v in vals:
            v = v.strip()
            if v and v not in seen:
                seen.add(v)
                out.append(v)
        return out

    def get_user_to_printers(self) -> dict:
        mapping = defaultdict(list)
        col = self.lists_df["UserHasPrinters"].dropna().astype(str).tolist()
        for entry in col:
            if "|" in entry:
                u, m = entry.split("|", 1)
            elif ":" in entry:
                u, m = entry.split(":", 1)
            else:
                continue
            u, m = u.strip(), m.strip()
            if u and m:
                mapping[u].append(m)
        return mapping

    def set_user_printers(self, user: str, models: list[str]):
        mask = self.lists_df["UserHasPrinters"].astype(str).str.startswith(f"{user}|") | \
               self.lists_df["UserHasPrinters"].astype(str).str.startswith(f"{user}:")
        self.lists_df = self.lists_df.loc[~mask].reset_index(drop=True)

        for m in models:
            new_row = {LISTS_COLUMNS[0]: None, LISTS_COLUMNS[1]: None, LISTS_COLUMNS[2]: f"{user}|{m}"}
            self.lists_df = pd.concat([self.lists_df, pd.DataFrame([new_row])], ignore_index=True)

        self.save()
        self.data_changed.emit()

    def add_user(self, user: str):
        if user and user not in self.get_usernames():
            new_row = {LISTS_COLUMNS[0]: None, LISTS_COLUMNS[1]: user, LISTS_COLUMNS[2]: None}
            self.lists_df = pd.concat([self.lists_df, pd.DataFrame([new_row])], ignore_index=True)
            self.save()
            self.data_changed.emit()

    def remove_user(self, user: str):
        if not user:
            return
        mask_users = self.lists_df["Usernames"].astype(str) == user
        mask_map = self.lists_df["UserHasPrinters"].astype(str).str.startswith(f"{user}|") | \
                   self.lists_df["UserHasPrinters"].astype(str).str.startswith(f"{user}:")
        self.lists_df = self.lists_df.loc[~(mask_users | mask_map)].reset_index(drop=True)
        self.save()
        self.data_changed.emit()

    def add_printer_model(self, model: str):
        if model and model not in self.get_printer_names():
            new_row = {LISTS_COLUMNS[0]: model, LISTS_COLUMNS[1]: None, LISTS_COLUMNS[2]: None}
            self.lists_df = pd.concat([self.lists_df, pd.DataFrame([new_row])], ignore_index=True)
            self.ensure_data_row_for_model(model)
            self.save()
            self.data_changed.emit()

    def remove_printer_model(self, model: str):
        if not model:
            return
        mask_models = self.lists_df["PrinterNames"].astype(str) == model
        mask_map = self.lists_df["UserHasPrinters"].astype(str).str.endswith(f"|{model}") | \
                   self.lists_df["UserHasPrinters"].astype(str).str.endswith(f":{model}")
        self.lists_df = self.lists_df.loc[~(mask_models | mask_map)].reset_index(drop=True)
        self.save()
        self.data_changed.emit()

    # ---------- DATA (Sheet 1) ----------

    def ensure_data_row_for_model(self, model: str):
        if (self.data_df["Model"].astype(str) == model).any():
            return
        new_row = {
            "Row": len(self.data_df) + 1,
            "Model": model,
            "New in storage": 0,
            "Storage": 0,
            "Repair": 0,
            "Total": 0,
            "User": ""
        }
        self.data_df = pd.concat([self.data_df, pd.DataFrame([new_row])], ignore_index=True)
        self.sync_totals(save=False)

    def sync_totals(self, save=True):
        if not self.data_df.empty:
            self.data_df["Total"] = (
                self.data_df["New in storage"].fillna(0).astype(int) +
                self.data_df["Storage"].fillna(0).astype(int) +
                self.data_df["Repair"].fillna(0).astype(int)
            ).astype(int)

        if save:
            self.save()
            self.data_changed.emit()

    def _row_index_by_model(self, model: str):
        matches = self.data_df.index[self.data_df["Model"].astype(str) == str(model)].tolist()
        return matches[0] if matches else None

    def adjust_storage(self, model: str, delta: int):
        idx = self._row_index_by_model(model)
        if idx is None:
            return False
        current = int(self.data_df.at[idx, "Storage"])
        new_val = current + delta
        if new_val < 0:
            return False
        self.data_df.at[idx, "Storage"] = new_val
        self.sync_totals(save=False)
        return True

    def adjust_repair(self, model: str, delta: int):
        idx = self._row_index_by_model(model)
        if idx is None:
            return False
        current = int(self.data_df.at[idx, "Repair"])
        new_val = current + delta
        if new_val < 0:
            return False
        self.data_df.at[idx, "Repair"] = new_val
        self.sync_totals(save=False)
        return True

    def set_user_on_data_row(self, model: str, user: str):
        idx = self._row_index_by_model(model)
        if idx is not None:
            self.data_df.at[idx, "User"] = user

    # ---------- LOGS (Sheet 2) ----------

    def add_log(self, jalali_date: str, event: str, user: str, model: str, qty: int, notes: str = ""):
        new_row = {
            LOG_COLUMNS[0]: jalali_date,
            LOG_COLUMNS[1]: event,
            LOG_COLUMNS[2]: user,
            LOG_COLUMNS[3]: model,
            LOG_COLUMNS[4]: int(qty),
            LOG_COLUMNS[5]: notes
        }
        self.logs_df = pd.concat([self.logs_df, pd.DataFrame([new_row])], ignore_index=True)

    # ---------- TOOLTIP AGGREGATION ----------

    def aggregate_user_model_info(self, model: str):
        mapping = self.get_user_to_printers()
        counts = {}
        for u, plist in mapping.items():
            c = sum(1 for p in plist if p == model)
            if c:
                counts[u] = c

        relevant = self.logs_df[
            (self.logs_df["Model"].astype(str) == model) &
            (self.logs_df["Event"].isin(["Issued", "IssuedReplacement"]))
        ]
        dates_by_user = defaultdict(list)
        for _, row in relevant.iterrows():
            u = str(row["User"]) if not pd.isna(row["User"]) else ""
            d = str(row["Date (Jalali)"]) if not pd.isna(row["Date (Jalali)"]) else ""
            q = int(row["Quantity"]) if not pd.isna(row["Quantity"]) else 1
            for _ in range(max(1, q)):
                if u and d:
                    dates_by_user[u].append(d)
        for u in dates_by_user:
            dates_by_user[u].reverse()

        return counts, dates_by_user
# =========================
# Qt Models / Views
# =========================

class PandasTableModel(QtCore.QAbstractTableModel):
    def __init__(self, df: pd.DataFrame, editable_columns=None, parent=None):
        super().__init__(parent)
        self._df = df
        self._editable = set(editable_columns or [])

    def rowCount(self, parent=QtCore.QModelIndex()):
        return 0 if parent.isValid() else len(self._df)

    def columnCount(self, parent=QtCore.QModelIndex()):
        return 0 if parent.isValid() else len(self._df.columns)

    def data(self, index, role=QtCore.Qt.DisplayRole):
        if not index.isValid():
            return None
        value = self._df.iat[index.row(), index.column()]
        if role in (QtCore.Qt.DisplayRole, QtCore.Qt.EditRole):
            return "" if pd.isna(value) else str(value)
        if role == QtCore.Qt.TextAlignmentRole:
            if self._df.columns[index.column()] in ["Row", "New in storage", "Storage", "Repair", "Total", "Quantity"]:
                return QtCore.Qt.AlignCenter
        return None

    def headerData(self, section, orientation, role=QtCore.Qt.DisplayRole):
        if role != QtCore.Qt.DisplayRole:
            return None
        if orientation == QtCore.Qt.Horizontal:
            return str(self._df.columns[section])
        else:
            return str(section + 1)

    def flags(self, index):
        base = QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled
        col_name = self._df.columns[index.column()]
        if col_name in self._editable:
            return base | QtCore.Qt.ItemIsEditable
        return base

    def setData(self, index, value, role=QtCore.Qt.EditRole):
        if role != QtCore.Qt.EditRole or not index.isValid():
            return False
        col_name = self._df.columns[index.column()]
        if col_name not in self._editable:
            return False
        # type coercion
        if col_name in ["New in storage", "Storage", "Repair", "Total", "Row", "Quantity"]:
            try:
                value = int(str(value).strip())
            except Exception:
                return False
        self._df.iat[index.row(), index.column()] = value
        self.dataChanged.emit(index, index, [QtCore.Qt.DisplayRole, QtCore.Qt.EditRole])
        return True

    def refresh(self):
        self.beginResetModel()
        self.endResetModel()


class HoverTableView(QtWidgets.QTableView):
    """
    Custom QTableView that shows tooltips with per-model aggregation when hovering Model/User columns.
    """
    def __init__(self, data_mgr: DataManager, parent=None):
        super().__init__(parent)
        self.data_mgr = data_mgr
        self.setMouseTracking(True)

    def viewportEvent(self, event: QtCore.QEvent) -> bool:
        if event.type() == QtCore.QEvent.ToolTip:
            help_event = QtGui.QHelpEvent(event)
            index = self.indexAt(help_event.pos())
            if index.isValid():
                model_col = self.model()._df.columns[index.column()]
                if model_col in ("Model", "User"):
                    # determine model of this row
                    row_model = self.model()._df.at[index.row(), "Model"]
                    counts, dates_by_user = self.data_mgr.aggregate_user_model_info(str(row_model))
                    if counts:
                        parts = []
                        for u, cnt in sorted(counts.items(), key=lambda x: (-x[1], x[0].lower())):
                            dates = ", ".join(dates_by_user.get(u, [])) or "—"
                            parts.append(f"• {u} — {cnt} unit(s)\n   dates: {dates}")
                        QtWidgets.QToolTip.showText(help_event.globalPos(),
                                                    f"Users of {row_model}:\n" + "\n".join(parts),
                                                    self)
                        return True
            QtWidgets.QToolTip.hideText()
            event.ignore()
        return super().viewportEvent(event)


# =========================
# Dialogs
# =========================

class EditDataRowDialog(QtWidgets.QDialog):
    def __init__(self, row_data: dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edit Data Row")
        self.setModal(True)
        form = QtWidgets.QFormLayout(self)

        self.le_row = QtWidgets.QLineEdit(str(row_data.get("Row", ""))); self.le_row.setReadOnly(True)
        self.le_model = QtWidgets.QLineEdit(str(row_data.get("Model", "")))
        self.sb_new = QtWidgets.QSpinBox(); self.sb_new.setRange(0, 10**9); self.sb_new.setValue(int(row_data.get("New in storage", 0)))
        self.sb_storage = QtWidgets.QSpinBox(); self.sb_storage.setRange(0, 10**9); self.sb_storage.setValue(int(row_data.get("Storage", 0)))
        self.sb_repair = QtWidgets.QSpinBox(); self.sb_repair.setRange(0, 10**9); self.sb_repair.setValue(int(row_data.get("Repair", 0)))
        self.le_total = QtWidgets.QLineEdit(); self.le_total.setReadOnly(True)
        self.le_user = QtWidgets.QLineEdit(str(row_data.get("User", "")))

        def refresh_total():
            t = self.sb_new.value() + self.sb_storage.value() + self.sb_repair.value()
            self.le_total.setText(str(t))

        for w in (self.sb_new, self.sb_storage, self.sb_repair):
            w.valueChanged.connect(refresh_total)
        refresh_total()

        form.addRow("Row:", self.le_row)
        form.addRow("Model:", self.le_model)
        form.addRow("New in storage:", self.sb_new)
        form.addRow("Storage:", self.sb_storage)
        form.addRow("Repair:", self.sb_repair)
        form.addRow("Total (auto):", self.le_total)
        form.addRow("User:", self.le_user)

        buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        form.addRow(buttons)

    def values(self):
        return {
            "Row": int(self.le_row.text() or "0"),
            "Model": self.le_model.text().strip(),
            "New in storage": int(self.sb_new.value()),
            "Storage": int(self.sb_storage.value()),
            "Repair": int(self.sb_repair.value()),
            "User": self.le_user.text().strip()
        }


class ReplaceAndSendRepairDialog(QtWidgets.QDialog):
    """
    The primary (most important) flow:
      - Choose a user
      - Choose a printer that this user already has (from 3rd sheet mapping)
      - Choose/edit date (defaults to today's Jalali)
      On OK:
         Storage -= 1  (a replacement unit goes out)
         Repair  += 1  (the broken unit goes to repair)
         Logs record both actions.
    """
    def __init__(self, data_mgr: DataManager, parent=None):
        super().__init__(parent)
        self.data_mgr = data_mgr
        self.setWindowTitle("Replace & Send to Repair")
        self.setModal(True)

        layout = QtWidgets.QFormLayout(self)

        self.cb_user = QtWidgets.QComboBox()
        self.cb_user.addItems(self.data_mgr.get_usernames())

        self.cb_model = QtWidgets.QComboBox()

        self.le_date = QtWidgets.QLineEdit(today_jalali_str())
        self.le_date.setPlaceholderText("YYYY/MM/DD (Jalali)")

        self.lbl_stock = QtWidgets.QLabel("")

        layout.addRow("User:", self.cb_user)
        layout.addRow("Printer (user has):", self.cb_model)
        layout.addRow("Date (Jalali):", self.le_date)
        layout.addRow("Storage available:", self.lbl_stock)

        self.cb_user.currentTextChanged.connect(self._reload_models_for_user)
        self.cb_model.currentTextChanged.connect(self._update_stock_label)
        self._reload_models_for_user(self.cb_user.currentText())

        self.buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        self.buttons.accepted.connect(self._on_ok)
        self.buttons.rejected.connect(self.reject)
        layout.addRow(self.buttons)

        self.ok_payload = None

    def _reload_models_for_user(self, user: str):
        self.cb_model.clear()
        mapping = self.data_mgr.get_user_to_printers()
        models = sorted(set(mapping.get(user, [])))
        self.cb_model.addItems(models)
        self._update_stock_label()

    def _update_stock_label(self):
        model = self.cb_model.currentText()
        idx = self.data_mgr._row_index_by_model(model)
        if idx is None:
            self.lbl_stock.setText("—")
            return
        stock = int(self.data_mgr.data_df.at[idx, "Storage"])
        self.lbl_stock.setText(str(stock))

    def _on_ok(self):
        user = self.cb_user.currentText().strip()
        model = self.cb_model.currentText().strip()
        date_jalali = self.le_date.text().strip() or today_jalali_str()

        if not user or not model:
            QtWidgets.QMessageBox.warning(self, "Missing", "Select a user and a model.")
            return

        idx = self.data_mgr._row_index_by_model(model)
        if idx is None:
            QtWidgets.QMessageBox.warning(self, "Unknown model", f"Model '{model}' not found in Data.")
            return
        stock = int(self.data_mgr.data_df.at[idx, "Storage"])
        if stock <= 0:
            QtWidgets.QMessageBox.warning(self, "Out of stock", f"No units of '{model}' available in Storage.")
            return

        self.ok_payload = (user, model, date_jalali)
        self.accept()


class RepairReturnDialog(QtWidgets.QDialog):
    """
    Choose quantities of models currently in Repair to return to Storage.
    """
    def __init__(self, data_mgr: DataManager, parent=None):
        super().__init__(parent)
        self.data_mgr = data_mgr
        self.setWindowTitle("Repair Return")
        self.setModal(True)

        vbox = QtWidgets.QVBoxLayout(self)
        self.form = QtWidgets.QFormLayout()
        vbox.addLayout(self.form)

        self.spins = {}  # model -> spinbox

        # rows for each model with Repair > 0
        df = self.data_mgr.data_df
        rows = df[df["Repair"] > 0][["Model", "Repair"]]
        for _, r in rows.iterrows():
            model = str(r["Model"])
            rep_count = int(r["Repair"])
            w = QtWidgets.QWidget()
            h = QtWidgets.QHBoxLayout(w); h.setContentsMargins(0, 0, 0, 0)
            spin = QtWidgets.QSpinBox(); spin.setRange(0, rep_count); spin.setValue(0); spin.setMaximumWidth(120)
            plus = QtWidgets.QToolButton(); plus.setText("➕"); plus.clicked.connect(lambda _, s=spin: s.setValue(min(s.maximum(), s.value()+1)))
            info = QtWidgets.QLabel(f" in repair: {rep_count}")
            h.addWidget(spin); h.addWidget(plus); h.addWidget(info); h.addStretch(1)
            self.form.addRow(model + ":", w)
            self.spins[model] = spin

        self.buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        self.buttons.accepted.connect(self._on_ok)
        self.buttons.rejected.connect(self.reject)
        vbox.addWidget(self.buttons)

        self.ok_payload = {}  # model -> qty

    def _on_ok(self):
        sel = {m: s.value() for m, s in self.spins.items() if s.value() > 0}
        if not sel:
            QtWidgets.QMessageBox.information(self, "Nothing selected", "Choose at least one quantity to return.")
            return
        self.ok_payload = sel
        self.accept()


# =========================
# Main Window
# =========================

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self, data_mgr: DataManager):
        super().__init__()
        self.data_mgr = data_mgr
        self.setWindowTitle("Printer DB")
        self.resize(1200, 720)

        # Global style (Fusion + subtle modern QSS)
        QtWidgets.QApplication.setStyle("Fusion")
        self.setStyleSheet("""
            QMainWindow { background: #121212; }
            QTabWidget::pane { border: 1px solid #2c2c2c; }
            QTabBar::tab { background: #1e1e1e; color: #ddd; padding: 8px 16px; margin: 2px; border-radius: 6px; }
            QTabBar::tab:selected { background: #2b2b2b; }
            QTableView { background: #181818; alternate-background-color: #202020; color: #ddd; gridline-color: #333; border: 1px solid #2c2c2c; }
            QHeaderView::section { background: #1b1b1b; color: #ddd; padding: 6px; border: none; border-right: 1px solid #333; }
            QLineEdit, QComboBox, QSpinBox, QDateEdit, QTextEdit { background: #1b1b1b; color: #eee; border: 1px solid #333; border-radius: 6px; padding: 4px 6px; }
            QPushButton { background: #2a2a2a; color: #eee; border: 1px solid #3a3a3a; border-radius: 8px; padding: 8px 14px; }
            QPushButton:hover { background: #333; }
            QPushButton#primary { background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #0ea5e9, stop:1 #0369a1); color: white; border: none; font-weight: 600; }
            QPushButton#primary:hover { filter: brightness(115%); }
            QToolButton { background: #2a2a2a; color: #eee; border: 1px solid #3a3a3a; border-radius: 6px; padding: 4px 8px; }
            QToolTip { background-color: #2b2b2b; color: #fff; border: 1px solid #3a3a3a; }
            QGroupBox { border: 1px solid #2c2c2c; border-radius: 8px; margin-top: 16px; color: #ddd; }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 6px; }
        """)

        self.tabs = QtWidgets.QTabWidget()
        self.setCentralWidget(self.tabs)

        self._build_tab_data()
        self._build_tab_logs()
        self._build_tab_manage()

        self.data_mgr.data_changed.connect(self._refresh_views)

    # ----- Tab: Data -----

    def _build_tab_data(self):
        tab = QtWidgets.QWidget()
        self.tabs.addTab(tab, "Data")

        vbox = QtWidgets.QVBoxLayout(tab)

        # Buttons row
        buttons = QtWidgets.QHBoxLayout()
        self.btn_primary = QtWidgets.QPushButton("Replace & Send to Repair")
        self.btn_primary.setObjectName("primary")
        self.btn_primary.setIcon(self._std_icon(QtWidgets.QStyle.SP_BrowserReload))  # just to make it visually distinct
        self.btn_edit = QtWidgets.QPushButton("Edit Selected…")
        self.btn_repair = QtWidgets.QPushButton("Repair Return…")
        self.btn_save = QtWidgets.QPushButton("Save")

        buttons.addWidget(self.btn_primary, 2)
        buttons.addStretch(1)
        buttons.addWidget(self.btn_edit)
        buttons.addWidget(self.btn_repair)
        buttons.addWidget(self.btn_save)
        vbox.addLayout(buttons)

        # Table
        self.data_model = PandasTableModel(self.data_mgr.data_df, editable_columns=[])  # editing via dialog
        self.tbl_data = HoverTableView(self.data_mgr)
        self.tbl_data.setModel(self.data_model)
        self.tbl_data.setAlternatingRowColors(True)
        self.tbl_data.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.tbl_data.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.tbl_data.horizontalHeader().setStretchLastSection(True)
        self.tbl_data.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        vbox.addWidget(self.tbl_data, 1)

        # Connections
        self.btn_save.clicked.connect(self._save_all)
        self.btn_edit.clicked.connect(self._edit_selected)
        self.btn_primary.clicked.connect(self._do_replace_send_repair)
        self.btn_repair.clicked.connect(self._do_repair_return)

    def _edit_selected(self):
        idx = self.tbl_data.selectionModel().currentIndex()
        if not idx.isValid():
            QtWidgets.QMessageBox.information(self, "No selection", "Select a row to edit.")
            return
        r = idx.row()
        row_data = {c: self.data_mgr.data_df.at[r, c] for c in DATA_COLUMNS}
        dlg = EditDataRowDialog(row_data, self)
        if dlg.exec() == QtWidgets.QDialog.Accepted:
            vals = dlg.values()
            for k in ["Model", "New in storage", "Storage", "Repair", "User"]:
                self.data_mgr.data_df.at[r, k] = vals[k]
            self.data_mgr.sync_totals(save=True)

    def _do_replace_send_repair(self):
        dlg = ReplaceAndSendRepairDialog(self.data_mgr, self)
        if dlg.exec() != QtWidgets.QDialog.Accepted or dlg.ok_payload is None:
            return
        user, model, jalali_date = dlg.ok_payload

        # Apply counts
        ok1 = self.data_mgr.adjust_storage(model, -1)
        ok2 = self.data_mgr.adjust_repair(model, +1)
        if not (ok1 and ok2):
            QtWidgets.QMessageBox.warning(self, "Error", "Could not adjust counts (negative?).")
            return

        # Update user (visually on Data tab)
        self.data_mgr.set_user_on_data_row(model, user)

        # Log actions
        self.data_mgr.add_log(jalali_date, "IssuedReplacement", user, model, 1, notes="Replaced user unit; broken to repair")
        self.data_mgr.add_log(jalali_date, "ToRepair", user, model, 1, notes="Broken unit sent to repair")

        self.data_mgr.save()
        self._refresh_views()

    def _do_repair_return(self):
        dlg = RepairReturnDialog(self.data_mgr, self)
        if dlg.exec() != QtWidgets.QDialog.Accepted:
            return
        sel = dlg.ok_payload  # {model: qty}
        jalali_date = today_jalali_str()
        for model, qty in sel.items():
            if qty <= 0:
                continue
            ok_r = self.data_mgr.adjust_repair(model, -qty)
            ok_s = self.data_mgr.adjust_storage(model, +qty)
            if ok_r and ok_s:
                self.data_mgr.add_log(jalali_date, "RepairReturned", "", model, qty, notes="Returned from repair to storage")
        self.data_mgr.save()
        self._refresh_views()

    def _save_all(self):
        self.data_mgr.sync_totals(save=True)
        QtWidgets.QMessageBox.information(self, "Saved", "All changes saved to Excel.")

    # ----- Tab: Logs -----

    def _build_tab_logs(self):
        tab = QtWidgets.QWidget()
        self.tabs.addTab(tab, "Logs")
        vbox = QtWidgets.QVBoxLayout(tab)

        self.logs_model = PandasTableModel(self.data_mgr.logs_df, editable_columns=[])
        self.tbl_logs = QtWidgets.QTableView()
        self.tbl_logs.setModel(self.logs_model)
        self.tbl_logs.setAlternatingRowColors(True)
        self.tbl_logs.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.tbl_logs.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.tbl_logs.horizontalHeader().setStretchLastSection(True)
        vbox.addWidget(self.tbl_logs, 1)

        # Quick export button
        btn_export = QtWidgets.QPushButton("Export Logs to CSV…")
        btn_export.clicked.connect(self._export_logs_csv)
        vbox.addWidget(btn_export, 0, alignment=QtCore.Qt.AlignRight)

    def _export_logs_csv(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Export Logs", "", "CSV Files (*.csv)")
        if not path:
            return
        try:
            self.data_mgr.logs_df.to_csv(path, index=False, encoding="utf-8-sig")
            QtWidgets.QMessageBox.information(self, "Exported", f"Logs exported to:\n{path}")
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error", f"Failed to export:\n{e}")

    # ----- Tab: Manage (Users & Printers) -----

    def _build_tab_manage(self):
        tab = QtWidgets.QWidget()
        self.tabs.addTab(tab, "Manage (users & printers)")
        grid = QtWidgets.QGridLayout(tab)

        # Group: Users
        grp_users = QtWidgets.QGroupBox("Define user & assign printers")
        grid.addWidget(grp_users, 0, 0)
        u_grid = QtWidgets.QGridLayout(grp_users)

        self.cb_users = QtWidgets.QComboBox(); self.cb_users.addItems(self.data_mgr.get_usernames())
        btn_rm_user = QtWidgets.QPushButton("Remove user")
        btn_rm_user.clicked.connect(self._remove_user)

        self.le_new_user = QtWidgets.QLineEdit(); self.le_new_user.setPlaceholderText("new username")
        btn_add_user = QtWidgets.QPushButton("Add user")
        btn_add_user.clicked.connect(self._add_user)

        u_grid.addWidget(QtWidgets.QLabel("Existing users:"), 0, 0)
        u_grid.addWidget(self.cb_users, 0, 1)
        u_grid.addWidget(btn_rm_user, 0, 2)
        u_grid.addWidget(QtWidgets.QLabel("Add user:"), 1, 0)
        u_grid.addWidget(self.le_new_user, 1, 1)
        u_grid.addWidget(btn_add_user, 1, 2)

        # Assign printers to user (only affects 3rd sheet & dropdowns)
        grp_assign = QtWidgets.QGroupBox("User ↔ Printers mapping (Sheet 3 only)")
        grid.addWidget(grp_assign, 1, 0)
        a_grid = QtWidgets.QGridLayout(grp_assign)

        self.cb_user_for_map = QtWidgets.QComboBox(); self.cb_user_for_map.addItems(self.data_mgr.get_usernames())
        self.cb_model_for_map = QtWidgets.QComboBox(); self.cb_model_for_map.addItems(self.data_mgr.get_printer_names())
        btn_add_map = QtWidgets.QPushButton("Add mapping")
        btn_add_map.clicked.connect(self._add_mapping)
        btn_rm_map = QtWidgets.QPushButton("Remove selected mapping")
        btn_rm_map.clicked.connect(self._remove_selected_mapping)

        self.lst_user_models = QtWidgets.QListWidget()
        self.cb_user_for_map.currentTextChanged.connect(self._reload_user_mappings)
        self._reload_user_mappings(self.cb_user_for_map.currentText())

        a_grid.addWidget(QtWidgets.QLabel("User:"), 0, 0)
        a_grid.addWidget(self.cb_user_for_map, 0, 1)
        a_grid.addWidget(QtWidgets.QLabel("Printer name:"), 1, 0)
        a_grid.addWidget(self.cb_model_for_map, 1, 1)
        a_grid.addWidget(btn_add_map, 1, 2)
        a_grid.addWidget(QtWidgets.QLabel("This user's printers:"), 2, 0, 1, 3)
        a_grid.addWidget(self.lst_user_models, 3, 0, 1, 3)
        a_grid.addWidget(btn_rm_map, 4, 0, 1, 3)

        # Group: Printers
        grp_printers = QtWidgets.QGroupBox("Define / remove printer names")
        grid.addWidget(grp_printers, 0, 1, 2, 1)
        p_grid = QtWidgets.QGridLayout(grp_printers)

        self.lst_models = QtWidgets.QListWidget()
        self._reload_models_list()

        self.le_new_model = QtWidgets.QLineEdit(); self.le_new_model.setPlaceholderText("new printer model")
        btn_add_model = QtWidgets.QPushButton("Add model")
        btn_add_model.clicked.connect(self._add_model)

        btn_rm_model = QtWidgets.QPushButton("Remove selected model(s)")
        btn_rm_model.clicked.connect(self._remove_models)

        p_grid.addWidget(QtWidgets.QLabel("Models (dropdown source):"), 0, 0, 1, 3)
        p_grid.addWidget(self.lst_models, 1, 0, 1, 3)
        p_grid.addWidget(self.le_new_model, 2, 0, 1, 2)
        p_grid.addWidget(btn_add_model, 2, 2)
        p_grid.addWidget(btn_rm_model, 3, 0, 1, 3)

        # Save button
        btn_save = QtWidgets.QPushButton("Save Lists")
        btn_save.clicked.connect(self._save_all)
        grid.addWidget(btn_save, 2, 1, alignment=QtCore.Qt.AlignRight)

    # Manage: Users
    def _add_user(self):
        user = self.le_new_user.text().strip()
        if not user:
            return
        self.data_mgr.add_user(user)
        self.le_new_user.clear()
        self._refresh_lists_controls()

    def _remove_user(self):
        user = self.cb_users.currentText().strip()
        if not user:
            return
        if QtWidgets.QMessageBox.question(self, "Confirm", f"Remove user '{user}'?") == QtWidgets.QMessageBox.Yes:
            self.data_mgr.remove_user(user)
            self._refresh_lists_controls()

    # Manage: Mapping
    def _add_mapping(self):
        user = self.cb_user_for_map.currentText().strip()
        model = self.cb_model_for_map.currentText().strip()
        if not user or not model:
            return
        cur = self.data_mgr.get_user_to_printers()
        models = cur.get(user, [])
        models.append(model)
        self.data_mgr.set_user_printers(user, models)
        self._reload_user_mappings(user)

    def _remove_selected_mapping(self):
        items = self.lst_user_models.selectedItems()
        if not items:
            return
        user = self.cb_user_for_map.currentText().strip()
        cur = self.data_mgr.get_user_to_printers()
        models = cur.get(user, [])
        for it in items:
            model = it.text().split("  ")[0]
            # remove one occurrence
            if model in models:
                models.remove(model)
        self.data_mgr.set_user_printers(user, models)
        self._reload_user_mappings(user)

    def _reload_user_mappings(self, user: str):
        self.lst_user_models.clear()
        mapping = self.data_mgr.get_user_to_printers()
        models = mapping.get(user, [])
        # show counts
        cnt = Counter(models)
        for m, n in sorted(cnt.items(), key=lambda x: (x[0].lower())):
            self.lst_user_models.addItem(f"{m}  ×{n}")

    # Manage: Printers
    def _add_model(self):
        model = self.le_new_model.text().strip()
        if not model:
            return
        self.data_mgr.add_printer_model(model)
        self.le_new_model.clear()
        self._reload_models_list()
        self._refresh_lists_controls()

    def _remove_models(self):
        items = self.lst_models.selectedItems()
        if not items:
            return
        names = [it.text() for it in items]
        if QtWidgets.QMessageBox.question(self, "Confirm", f"Remove selected model(s)?\n\n" + "\n".join(names)) != QtWidgets.QMessageBox.Yes:
            return
        for m in names:
            self.data_mgr.remove_printer_model(m)
        self._reload_models_list()
        self._refresh_lists_controls()

    def _reload_models_list(self):
        self.lst_models.clear()
        for m in self.data_mgr.get_printer_names():
            self.lst_models.addItem(m)

    # ----- Utilities -----

    def _std_icon(self, std):
        return self.style().standardIcon(std)

    def _refresh_lists_controls(self):
        # users
        users = self.data_mgr.get_usernames()
        for cb in (self.cb_users, self.cb_user_for_map):
            cur = cb.currentText()
            cb.blockSignals(True)
            cb.clear()
            cb.addItems(users)
            if cur in users:
                cb.setCurrentText(cur)
            cb.blockSignals(False)
        # models
        models = self.data_mgr.get_printer_names()
        curm = self.cb_model_for_map.currentText()
        self.cb_model_for_map.blockSignals(True)
        self.cb_model_for_map.clear()
        self.cb_model_for_map.addItems(models)
        if curm in models:
            self.cb_model_for_map.setCurrentText(curm)
        self.cb_model_for_map.blockSignals(False)
        # lists
        self._reload_models_list()
        self._reload_user_mappings(self.cb_user_for_map.currentText())

    def _refresh_views(self):
        # recompute totals in memory only (NO SAVE here)
        self.data_mgr.sync_totals(save=False)

        # refresh table models
        self.data_model.refresh()
        self.logs_model.refresh()

        # refresh dropdowns / lists in Manage tab
        self._refresh_lists_controls()


    # Close event: save
    def closeEvent(self, e: QtGui.QCloseEvent):
        try:
            self.data_mgr.sync_totals(save=True)
        except Exception:
            pass
        super().closeEvent(e)



# =========================
# Entry Point
# =========================

def main():
    app = QtWidgets.QApplication(sys.argv)
    dm = DataManager(EXCEL_PATH)
    w = MainWindow(dm)
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
